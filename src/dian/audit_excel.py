"""Exportación del paquete de auditoría a Excel (.xlsx)."""
import json
import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return str(value)
    return json.dumps(value, ensure_ascii=False)


def _flatten_dict(obj: dict[str, Any], prefix: str = "") -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = []
    for k, v in obj.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            rows.extend(_flatten_dict(v, key))
        else:
            rows.append((key, _cell_str(v)))
    return rows


def audit_pack_to_xlsx_bytes(pack: dict[str, Any]) -> bytes:
    """Genera un libro con hoja de datos de factura/perfil y hoja de eventos."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Datos"
    header_font = Font(bold=True)
    ws.cell(1, 1, "Campo").font = header_font
    ws.cell(1, 2, "Valor").font = header_font

    row = 2
    for k in ("generated_at", "schema_version", "organization_id"):
        if k in pack:
            ws.cell(row, 1, k)
            ws.cell(row, 2, _cell_str(pack[k]))
            row += 1

    inv = pack.get("invoice")
    if isinstance(inv, dict):
        ws.cell(row, 1, "— Factura —").font = header_font
        row += 1
        for field, val in _flatten_dict(inv, "invoice"):
            ws.cell(row, 1, field)
            ws.cell(row, 2, val)
            row += 1

    fp = pack.get("fiscal_profile")
    if isinstance(fp, dict):
        ws.cell(row, 1, "— Perfil fiscal —").font = header_font
        row += 1
        for field, val in _flatten_dict(fp, "perfil_fiscal"):
            ws.cell(row, 1, field)
            ws.cell(row, 2, val)
            row += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 52

    ev_sheet = wb.create_sheet("Eventos")
    ev_sheet.cell(1, 1, "ID").font = header_font
    ev_sheet.cell(1, 2, "Tipo").font = header_font
    ev_sheet.cell(1, 3, "Actor (user id)").font = header_font
    ev_sheet.cell(1, 4, "Fecha").font = header_font
    ev_sheet.cell(1, 5, "Payload (JSON)").font = header_font
    er = 2
    for e in pack.get("events") or []:
        if not isinstance(e, dict):
            continue
        payload = e.get("payload")
        payload_s = json.dumps(payload, ensure_ascii=False) if payload is not None else ""
        ev_sheet.cell(er, 1, e.get("id"))
        ev_sheet.cell(er, 2, e.get("event_type"))
        ev_sheet.cell(er, 3, e.get("actor_user_id") if e.get("actor_user_id") is not None else "")
        ev_sheet.cell(er, 4, e.get("created_at"))
        ev_sheet.cell(er, 5, payload_s)
        er += 1
    ev_sheet.column_dimensions["A"].width = 10
    ev_sheet.column_dimensions["B"].width = 22
    ev_sheet.column_dimensions["C"].width = 14
    ev_sheet.column_dimensions["D"].width = 24
    ev_sheet.column_dimensions["E"].width = 60

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def safe_audit_filename(invoice_number: str) -> str:
    """Nombre seguro para Content-Disposition."""
    s = re.sub(r"[^\w.\-]+", "_", invoice_number.strip(), flags=re.UNICODE)
    return (s or "factura")[:120]
