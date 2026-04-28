"""Generación de archivos XLSX y PDF para reportes de facturas."""
from __future__ import annotations

import io
import re
from html import escape
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image as RLImage
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from src.models import Invoice
from src.reporting.chart_images import dashboard_figure_png
from src.schemas import DashboardStatsOut


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    if dt.tzinfo is not None:
        dt = dt.astimezone(tz=None).replace(tzinfo=None)
    return dt.strftime("%Y-%m-%d %H:%M")


def _fmt_money(v: Decimal) -> str:
    return f"{v:,.2f}"


def _safe_filename_part(name: str) -> str:
    return re.sub(r"[^\w\-]+", "_", name, flags=re.UNICODE)[:80] or "reporte"


def _safe_xlsx_text(value: str | None) -> str:
    txt = (value or "").strip()
    if txt.startswith(("=", "+", "-", "@")):
        return f"'{txt}"
    return txt


def build_invoices_xlsx_bytes(
    invoices: list[Invoice],
    sheet_title: str = "Facturas",
    dashboard_stats: DashboardStatsOut | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] if sheet_title else "Facturas"
    headers = [
        "ID",
        "Número",
        "Proveedor",
        "Monto",
        "Estado",
        "Moneda",
        "Emisión",
        "Vencimiento",
        "Creado",
    ]
    bold = Font(bold=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
    for row_idx, inv in enumerate(invoices, 2):
        ws.cell(row=row_idx, column=1, value=inv.id)
        ws.cell(row=row_idx, column=2, value=_safe_xlsx_text(inv.invoice_number))
        ws.cell(row=row_idx, column=3, value=_safe_xlsx_text(inv.supplier))
        ws.cell(row=row_idx, column=4, value=float(inv.amount))
        ws.cell(row=row_idx, column=5, value=inv.status.value)
        ws.cell(row=row_idx, column=6, value=inv.currency)
        ws.cell(row=row_idx, column=7, value=_fmt_dt(inv.issue_date))
        ws.cell(row=row_idx, column=8, value=_fmt_dt(inv.due_date))
        ws.cell(row=row_idx, column=9, value=_fmt_dt(inv.created_at))

    if dashboard_stats is not None:
        png = dashboard_figure_png(dashboard_stats)
        ws_ch = wb.create_sheet("Gráficos")
        ws_ch["A1"] = "Dashboard (misma vista que la aplicación)"
        ws_ch["A1"].font = Font(bold=True, size=14)
        xl_img = XLImage(io.BytesIO(png))
        xl_img.width = 900
        xl_img.height = 675
        ws_ch.add_image(xl_img, "A3")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_invoices_pdf_bytes(
    invoices: list[Invoice],
    title: str,
    subtitle: str | None = None,
    dashboard_stats: DashboardStatsOut | None = None,
) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(escape(title), styles["Title"]),
    ]
    if subtitle:
        story.append(Paragraph(escape(subtitle), styles["Normal"]))
    story.append(Spacer(1, 14))

    if dashboard_stats is not None:
        png = dashboard_figure_png(dashboard_stats)
        # BytesIO tiene .read(); ImageReader no, y RL Image trata eso como ruta (splitext falla).
        story.append(
            RLImage(
                io.BytesIO(png),
                width=7.2 * inch,
                height=5.4 * inch,
            )
        )
        story.append(Spacer(1, 16))
        story.append(Paragraph(f"<b>{escape('Detalle de facturas')}</b>", styles["Normal"]))
        story.append(Spacer(1, 8))

    data = [
        [
            "ID",
            "Número",
            "Proveedor",
            "Monto",
            "Estado",
            "Moneda",
            "Emisión",
            "Vencimiento",
        ]
    ]
    for inv in invoices:
        data.append(
            [
                str(inv.id),
                inv.invoice_number[:24],
                (inv.supplier or "")[:40],
                _fmt_money(inv.amount),
                inv.status.value,
                inv.currency,
                _fmt_dt(inv.issue_date)[:10],
                _fmt_dt(inv.due_date)[:10] if inv.due_date else "",
            ]
        )

    col_widths = [28, 72, 200, 72, 70, 44, 72, 72]
    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0e7490")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
            ]
        )
    )
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def export_filename_prefix(organization_slug: str) -> str:
    return _safe_filename_part(organization_slug)
